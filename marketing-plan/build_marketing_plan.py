# -*- coding: utf-8 -*-
"""Generate Sky Lines annual marketing plan workbook (Excel).

Run: python3 build_marketing_plan.py
Output: Sky_Lines_Marketing_Plan_2026.xlsx (next to this script)
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Styles ----------
TITLE_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FILL = PatternFill("solid", start_color="6366F1")
ACCENT_FILL = PatternFill("solid", start_color="FEF3C7")
GREEN_FILL = PatternFill("solid", start_color="D1FAE5")
GRAY_FILL = PatternFill("solid", start_color="F3F4F6")

TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=15)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(name="Arial", bold=True, size=11)
BODY_FONT = Font(name="Arial", size=11)
INPUT_FONT = Font(name="Arial", color="0000FF", size=11)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
RIGHT_TOP = Alignment(horizontal="right", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="D1D5DB")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

wb = Workbook()


def title_row(ws, text, span):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = CENTER
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    ws.row_dimensions[1].height = 36


def header_cells(ws, headers, row=3):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 1: OVERVIEW
# ============================================================
ws = wb.active
ws.title = "Overview"
ws.sheet_view.rightToLeft = True
title_row(ws, "خطة التسويق السنوية — Sky Lines Group (2026/2027)", 2)

overview = [
    ["العنصر", "القيمة"],
    ["الشركة", "Sky Lines Group | SkyLines Developments"],
    ["النشاط", "تطوير + استثمار + تسويق عقاري — بني سويف، شرق النيل"],
    ["المشروع النشط", "Sky Villas M7 — 5 وحدات معروضة (سكني/إداري/تجاري)"],
    ["الاستلام المتوقع", "1 يوليو 2027"],
    ["الهدف الاستراتيجي", "بناء براند + توليد ليدات مؤهلة مستمر"],
    ["المدى الزمني", "12 شهر (يونيو 2026 → مايو 2027)"],
    ["الميزانية الشهرية", "15,000 – 30,000 ج (متوسط ~22,000)"],
    ["الميزانية السنوية (مرجعي)", "≈ 264,000 ج"],
    ["القنوات الرئيسية", "Meta Ads + Organic + TikTok + WhatsApp/Bot"],
    ["العمود الفقري", "المساعد الذكي (ماسنجر + واتساب) — تأهيل لحظي 24/7"],
    ["ليدات مستهدفة/سنة", "8,000 – 12,000 ليد"],
    ["تكلفة الليد المؤهل", "≤ 35 ج (CPQL)"],
    ["نمو المتابعين/سنة", "+15,000 (FB + IG + TikTok)"],
    ["تقرير يومي", "تليجرام 9 م (إنفاق + ليدات + تكلفة الليد)"],
    ["العروض المعتمدة", "مقدم 40%/15 شهر · 60%/12 شهر · كاش (بدون فوايد)"],
]
for r, (k, v) in enumerate(overview, start=3):
    a = ws.cell(row=r, column=1, value=k)
    b = ws.cell(row=r, column=2, value=v)
    a.border = b.border = BORDER
    a.alignment = CENTER
    b.alignment = RIGHT
    if r == 3:
        a.fill = b.fill = HEADER_FILL
        a.font = b.font = HEADER_FONT
        b.alignment = CENTER
    else:
        a.fill = ACCENT_FILL
        a.font = LABEL_FONT
        b.font = BODY_FONT
set_widths(ws, [32, 62])

# Note row
note_r = 3 + len(overview) + 1
ws.cell(row=note_r, column=1,
        value="⛔ ممنوع ذكر عروض منتهية (مثل: الشقة عليك والخدمات علينا / خصم 200 ألف). يُذكر فقط العروض المعتمدة أعلاه.")
ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=2)
ws.cell(row=note_r, column=1).fill = PatternFill("solid", start_color="FEE2E2")
ws.cell(row=note_r, column=1).font = LABEL_FONT
ws.cell(row=note_r, column=1).alignment = RIGHT
ws.row_dimensions[note_r].height = 30

# ============================================================
# SHEET 2: BUDGET (annual + monthly logic)
# ============================================================
ws = wb.create_sheet("Budget")
ws.sheet_view.rightToLeft = True
title_row(ws, "الميزانية — توزيع شهري وسنوي", 5)
header_cells(ws, ["#", "البند", "شهري (ج)", "سنوي (ج)", "النسبة %"])

budget = [
    [1, "إعلانات Meta (FB + Instagram)", 13000],
    [2, "إنتاج محتوى (تصميم/كتابة/مونتاج)", 3500],
    [3, "تصوير درون/فوتو (موزّع)", 2500],
    [4, "TikTok / Boosting / Micro-influencer", 1500],
    [5, "أدوات وبرمجيات (استضافة/جدولة/CRM)", 1000],
    [6, "اختبار واحتياطي", 500],
]
start = 4
for i, row in enumerate(budget):
    r = start + i
    ws.cell(row=r, column=1, value=row[0])
    ws.cell(row=r, column=2, value=row[1])
    ws.cell(row=r, column=3, value=row[2]).font = INPUT_FONT
    ws.cell(row=r, column=4, value=f"=C{r}*12")
    ws.cell(row=r, column=5, value=f"=C{r}/C{start+len(budget)}")
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = CENTER if c != 2 else RIGHT
        if c not in (3,):
            cell.font = BODY_FONT
    ws.cell(row=r, column=3).number_format = "#,##0"
    ws.cell(row=r, column=4).number_format = "#,##0"
    ws.cell(row=r, column=5).number_format = "0%"

total_r = start + len(budget)
ws.cell(row=total_r, column=1, value="")
ws.cell(row=total_r, column=2, value="الإجمالي")
ws.cell(row=total_r, column=3, value=f"=SUM(C{start}:C{total_r-1})")
ws.cell(row=total_r, column=4, value=f"=SUM(D{start}:D{total_r-1})")
ws.cell(row=total_r, column=5, value="=C{0}/C{0}".format(total_r))
for c in range(1, 6):
    cell = ws.cell(row=total_r, column=c)
    cell.fill = ACCENT_FILL
    cell.border = BORDER
    cell.font = LABEL_FONT
    cell.alignment = CENTER if c != 2 else RIGHT
ws.cell(row=total_r, column=3).number_format = "#,##0"
ws.cell(row=total_r, column=4).number_format = "#,##0"
ws.cell(row=total_r, column=5).number_format = "0%"
set_widths(ws, [5, 40, 16, 16, 12])

# Monthly budget plan (12 months, adjustable)
mr = total_r + 3
ws.cell(row=mr, column=1, value="الخطة الشهرية (قابلة للتعديل حسب الموسم)").fill = TITLE_FILL
ws.cell(row=mr, column=1).font = HEADER_FONT
ws.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=5)
ws.cell(row=mr, column=1).alignment = CENTER
header_cells(ws, ["الشهر", "المرحلة", "ميزانية الشهر (ج)", "ملاحظة", ""], row=mr+1)
months = [
    ("الشهر 1 (يونيو)", "تأسيس", 18000, "اختبار + إطلاق"),
    ("الشهر 2 (يوليو)", "تأسيس", 20000, "موسم المغتربين"),
    ("الشهر 3 (أغسطس)", "تأسيس", 20000, "تحسين"),
    ("الشهر 4 (سبتمبر)", "توسّع", 24000, "Retargeting + Lookalike"),
    ("الشهر 5 (أكتوبر)", "توسّع", 24000, "TikTok بجدّية"),
    ("الشهر 6 (نوفمبر)", "توسّع", 26000, "شهادات العملاء"),
    ("الشهر 7 (ديسمبر)", "سلطة براند", 22000, "محتوى احترافي"),
    ("الشهر 8 (يناير)", "سلطة براند", 22000, "Google/YouTube تجريبي"),
    ("الشهر 9 (فبراير)", "سلطة براند", 24000, "تشويق Sky East"),
    ("الشهر 10 (مارس)", "حصاد", 26000, "رمضان — محتوى عاطفي"),
    ("الشهر 11 (أبريل)", "حصاد", 26000, "العيد + إغلاق"),
    ("الشهر 12 (مايو)", "حصاد", 22000, "مراجعة سنوية"),
]
hr = mr + 2
for i, m in enumerate(months):
    r = hr + i
    ws.cell(row=r, column=1, value=m[0])
    ws.cell(row=r, column=2, value=m[1])
    ws.cell(row=r, column=3, value=m[2]).font = INPUT_FONT
    ws.cell(row=r, column=3).number_format = "#,##0"
    ws.cell(row=r, column=4, value=m[3])
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = CENTER if c in (2, 3) else RIGHT
        if c != 3:
            cell.font = BODY_FONT
gr = hr + len(months)
ws.cell(row=gr, column=1, value="إجمالي السنة")
ws.cell(row=gr, column=2, value="")
ws.cell(row=gr, column=3, value=f"=SUM(C{hr}:C{gr-1})")
ws.cell(row=gr, column=3).number_format = "#,##0"
for c in range(1, 6):
    cell = ws.cell(row=gr, column=c)
    cell.fill = ACCENT_FILL
    cell.border = BORDER
    cell.font = LABEL_FONT
    cell.alignment = CENTER

# ============================================================
# SHEET 3: TEAM & ROLES
# ============================================================
ws = wb.create_sheet("Team_Roles")
ws.sheet_view.rightToLeft = True
title_row(ws, "أدوار الفريق والمسؤوليات", 4)
header_cells(ws, ["الدور", "المسؤوليات", "المخرجات الأسبوعية", "المسؤول (اسم)"])
roles = [
    ("مدير التسويق", "الاستراتيجية، الميزانية، التقارير، التنسيق مع الإدارة", "تقرير أسبوعي + قرارات ميزانية", ""),
    ("Media Buyer / أداء", "إعداد وتحسين حملات Meta، المتابعة اليومية", "تقرير إنفاق + ليدات يومي", ""),
    ("كاتب محتوى", "كابشن، نصوص إعلانات، سكربتات فيديو", "خطة محتوى الأسبوع + النصوص", ""),
    ("مصمم جرافيك", "كرياتيف، كاروسيل، الهوية البصرية", "3–5 تصاميم/أسبوع", ""),
    ("مصوّر / مونتير", "تصوير درون، ريلز، مونتاج تقدّم البناء", "تحديث بناء كل أسبوعين + ريلز", ""),
    ("Community Manager", "الرد على الكومنتات، التفاعل، تحويل المهتم", "تفاعل يومي + تحويل ليدات", ""),
    ("متابعة المبيعات", "استلام الليدات المؤهلة، الاتصال، المعاينة، الإغلاق", "تحديث حالة الليدات + معاينات", ""),
    ("المساعد الذكي (Bot)", "الرد الأول وتأهيل الليد لحظياً 24/7", "تأهيل تلقائي + تسجيل", "آلي"),
]
for i, row in enumerate(roles):
    r = 4 + i
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = RIGHT if c in (2, 3) else CENTER
        cell.font = LABEL_FONT if c == 1 else BODY_FONT
        if c == 1:
            cell.fill = ACCENT_FILL
        if c == 4:
            cell.font = INPUT_FONT
set_widths(ws, [22, 46, 34, 20])

# Meeting rhythm
mrr = 4 + len(roles) + 2
ws.cell(row=mrr, column=1, value="إيقاع الاجتماعات").fill = TITLE_FILL
ws.cell(row=mrr, column=1).font = HEADER_FONT
ws.merge_cells(start_row=mrr, start_column=1, end_row=mrr, end_column=4)
ws.cell(row=mrr, column=1).alignment = CENTER
header_cells(ws, ["الإيقاع", "المدة", "الهدف", "المشاركون"], row=mrr+1)
meetings = [
    ("يومي", "15 دقيقة", "مراجعة الإنفاق والليدات", "Media Buyer + المدير"),
    ("أسبوعي", "45 دقيقة", "الأداء + الكرياتيف الفائز + خطة الأسبوع", "الفريق كامل"),
    ("شهري", "60 دقيقة", "المؤشرات مقابل المستهدف + الميزانية", "الفريق + الإدارة"),
    ("ربع سنوي", "ساعتين", "مراجعة استراتيجية + أولويات الربع", "الفريق + الإدارة"),
]
for i, m in enumerate(meetings):
    r = mrr + 2 + i
    for c, v in enumerate(m, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = CENTER if c in (1, 2) else RIGHT
        cell.font = BODY_FONT

# ============================================================
# SHEET 4: CHANNELS
# ============================================================
ws = wb.create_sheet("Channels")
ws.sheet_view.rightToLeft = True
title_row(ws, "خطة القنوات التسويقية", 5)
header_cells(ws, ["القناة", "الدور", "الهدف", "الإيقاع/الميزانية", "المؤشر الرئيسي"])
channels = [
    ("Meta Ads (FB+IG)", "ماكينة الليدات الأساسية", "Messages + Lead Forms", "≈13,000 ج/شهر", "تكلفة الليد"),
    ("Organic (FB+IG)", "بناء البراند والثقة", "تقدّم بناء + تعليمي", "4–5 منشور/أسبوع", "Engagement + متابعين"),
    ("TikTok", "وصول واسع رخيص", "ريلز درون + شريحة أصغر", "3 فيديو/أسبوع", "Views + Reach"),
    ("WhatsApp + Bot", "التقاط وتأهيل الليد", "تأهيل لحظي 24/7", "آلي (موجود)", "ليدات مؤهلة"),
    ("الإحالة الأرضية", "ذراع بني سويف + لافتات", "تحويلات محلية", "حسب الفعالية", "إحالات/زيارات"),
    ("Email / SMS", "رعاية الليدات الباردة", "متابعة كل أسبوعين", "أداة جدولة", "معدل التحويل"),
    ("Google / YouTube", "بحث + فيديو (Q3/Q4)", "نية شراء عالية", "تجريبي لاحقاً", "تكلفة التحويل"),
]
for i, ch in enumerate(channels):
    r = 4 + i
    for c, v in enumerate(ch, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = RIGHT if c in (2, 3) else CENTER
        cell.font = LABEL_FONT if c == 1 else BODY_FONT
        if c == 1:
            cell.fill = ACCENT_FILL
set_widths(ws, [20, 26, 28, 22, 20])

# ============================================================
# SHEET 5: AUDIENCE / PERSONAS
# ============================================================
ws = wb.create_sheet("Audience")
ws.sheet_view.rightToLeft = True
title_row(ws, "الجمهور المستهدف والـ Personas", 2)
audience = [
    ("🔵 Persona 1 — أهل بني سويف (المقيم المحلي)", ""),
    ("العمر", "28–55"),
    ("الدافع", "سكن عائلي أو استثمار قريب يقدر يتابعه بنفسه"),
    ("الرسالة", "بيتك في قلب شرق النيل، بنظام سداد مريح"),
    ("الموقع", "بني سويف (40 كم) + المقيمين فيها"),
    ("الاهتمامات", "Real estate, Property, Investment, Home improvement"),
    ("القناة", "Meta Ads محلي + Organic + إحالة أرضية"),
    ("", ""),
    ("🟢 Persona 2 — المغترب البنّي سويفي (Gulf Expat)", ""),
    ("الدافع", "استثمار آمن بالجنيه في بلده، بيت العيلة، التقاعد"),
    ("الرسالة", "استثمر في بلدك من بعيد — وإحنا بنرتّب كل حاجة عن بُعد"),
    ("الموقع", "السعودية، الإمارات، الكويت، قطر، عُمان"),
    ("الاستهداف", "Expats — lived in Egypt + خرّيجي/مدارس بني سويف"),
    ("القناة", "Meta Ads خليج + WhatsApp + فيديو ثقة"),
    ("", ""),
    ("🟡 Persona 3 — المستثمر العقاري", ""),
    ("الدافع", "العائد، موقع تجاري ناصية، قناة إعادة بيع الشركة"),
    ("الرسالة", "موقع تجاري قوي + ذراع تسويق يساعدك تبيع وقت ما تحب"),
    ("الاهتمامات", "Real estate investing, Property investment, Egyptian real estate"),
    ("القناة", "Meta Ads (Interests) + Retargeting + LinkedIn لاحقاً"),
    ("", ""),
    ("🟣 Retargeting (جمهور دافئ)", ""),
    ("Custom Audience", "تفاعل البيدج 90 يوم + اللي بعتوا رسالة + Video Viewers 50%+"),
    ("Lookalike", "2% من قائمة العملاء/الليدات (وقت ما تتوفر)"),
    ("", ""),
    ("⚠️ التزام", "ممنوع وعد العميل بعائد مضمون أو ارتفاع أسعار/إيجارات مستقبلية"),
]
for i, (k, v) in enumerate(audience, start=3):
    a = ws.cell(row=i, column=1, value=k)
    b = ws.cell(row=i, column=2, value=v)
    a.border = b.border = BORDER
    a.alignment = RIGHT
    b.alignment = RIGHT
    if k.startswith(("🔵", "🟢", "🟡", "🟣")):
        a.fill = HEADER_FILL
        a.font = HEADER_FONT
        a.alignment = CENTER
        b.fill = ACCENT_FILL
    elif k.startswith("⚠️"):
        a.fill = PatternFill("solid", start_color="FEE2E2")
        a.font = LABEL_FONT
        b.fill = PatternFill("solid", start_color="FEE2E2")
        b.font = LABEL_FONT
    else:
        a.font = LABEL_FONT
        b.font = BODY_FONT
set_widths(ws, [38, 70])

# ============================================================
# SHEET 6: CONTENT CALENDAR
# ============================================================
ws = wb.create_sheet("Content_Calendar")
ws.sheet_view.rightToLeft = True
title_row(ws, "تقويم المحتوى — محاور شهرية وإيقاع أسبوعي", 4)

# Pillars block
ws.cell(row=3, column=1, value="ركائز المحتوى (5 محاور)").fill = HEADER_FILL
ws.cell(row=3, column=1).font = HEADER_FONT
ws.merge_cells("A3:D3")
ws.cell(row=3, column=1).alignment = CENTER
pillars = [
    ("1. تقدّم البناء", "صور/فيديو من الموقع كل أسبوعين → ثقة"),
    ("2. تعليمي", "شرح أنظمة السداد، نصائح استثمار، نمو بني سويف"),
    ("3. الوحدات والعروض", "عرض الوحدات المتاحة بشكل منظّم (المعتمد فقط)"),
    ("4. الثقة والمصداقية", "قصة الشركة، رؤيتها، شهادات العملاء"),
    ("5. الموقع والمجتمع", "معالم شرق النيل، أسلوب الحياة"),
]
for i, (k, v) in enumerate(pillars):
    r = 4 + i
    ws.cell(row=r, column=1, value=k).font = LABEL_FONT
    ws.cell(row=r, column=1).fill = ACCENT_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=1).alignment = RIGHT
    c = ws.cell(row=r, column=2, value=v)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c.font = BODY_FONT
    c.border = BORDER
    c.alignment = RIGHT

# Monthly themes
tr = 4 + len(pillars) + 1
ws.cell(row=tr, column=1, value="المحاور الشهرية").fill = HEADER_FILL
ws.cell(row=tr, column=1).font = HEADER_FONT
ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
ws.cell(row=tr, column=1).alignment = CENTER
header_cells(ws, ["الشهر", "المحور الرئيسي", "حملة/مناسبة", "نوع المحتوى الأبرز"], row=tr+1)
themes = [
    ("يونيو", "تعريف بالمشروع والموقع", "إطلاق", "فيديو درون + كاروسيل"),
    ("يوليو", "المغتربون العائدون", "صيف المعاينات", "فيديو ثقة + شهادات"),
    ("أغسطس", "أنظمة السداد المرنة", "تعليمي", "إنفوجرافيك + Reels"),
    ("سبتمبر", "تقدّم البناء", "ثقة", "تحديث موقع + Before/After"),
    ("أكتوبر", "الوحدات التجارية/الإدارية", "المستثمر", "كاروسيل عوائد الموقع"),
    ("نوفمبر", "شهادات أول المشترين", "Social Proof", "فيديو شهادات"),
    ("ديسمبر", "حصاد السنة + رؤية", "براند", "فيديو احترافي"),
    ("يناير", "بداية سنة استثمارية", "تعليمي", "نصائح + Lead magnet"),
    ("فبراير", "تشويق Sky East", "تشويق", "Teaser + قائمة اهتمام"),
    ("مارس", "رمضان — بيت العيلة", "موسمي عاطفي", "فيديو رمضاني"),
    ("أبريل", "العيد + إغلاق", "عروض موسمية", "عروض ضمن المعتمد"),
    ("مايو", "مراجعة + المشروع القادم", "إطلاق جديد", "إعلان رسمي"),
]
for i, t in enumerate(themes):
    r = tr + 2 + i
    for c, v in enumerate(t, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = CENTER if c == 1 else RIGHT
        cell.font = BODY_FONT

# Weekly cadence
wr = tr + 2 + len(themes) + 1
ws.cell(row=wr, column=1, value="الإيقاع الأسبوعي").fill = HEADER_FILL
ws.cell(row=wr, column=1).font = HEADER_FONT
ws.merge_cells(start_row=wr, start_column=1, end_row=wr, end_column=4)
ws.cell(row=wr, column=1).alignment = CENTER
header_cells(ws, ["المنصة", "العدد/الأسبوع", "النوع", "ملاحظة"], row=wr+1)
cadence = [
    ("Facebook + Instagram", "4–5 منشور + Stories يومية", "متنوّع حسب المحاور", "أساسي"),
    ("TikTok / Reels", "3 فيديو", "درون + لقطات سريعة", "وصول واسع"),
    ("تقدّم البناء", "كل أسبوعين", "Reel + ألبوم صور", "ثقة"),
    ("Email / SMS", "كل أسبوعين", "رعاية ليدات", "غير المغلقين"),
]
for i, cd in enumerate(cadence):
    r = wr + 2 + i
    for c, v in enumerate(cd, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = RIGHT if c in (3, 4) else CENTER
        cell.font = BODY_FONT
set_widths(ws, [22, 26, 30, 22])

# ============================================================
# SHEET 7: KPIs (monthly targets)
# ============================================================
ws = wb.create_sheet("KPIs")
ws.sheet_view.rightToLeft = True
title_row(ws, "المؤشرات والمستهدفات الشهرية", 4)
header_cells(ws, ["المؤشر", "المستهدف الشهري", "المستهدف السنوي", "ملاحظة"])
kpis = [
    ("Reach", 100000, "=B4*12", "وصول"),
    ("Impressions", 400000, "=B5*12", "ظهور"),
    ("Engagement Rate %", "3%", "—", "تفاعل"),
    ("Messages / Conversations", 800, "=B7*12", "هدف أساسي"),
    ("Leads (فورم + رسالة)", 850, "=B8*12", "توليد"),
    ("Cost per Lead (ج)", 12, "—", "Hard cap"),
    ("Qualified Leads (من البوت)", 220, "=B10*12", "≈25–30%"),
    ("Cost per Qualified Lead (ج)", 35, "—", "CPQL"),
    ("معاينات / اجتماعات", 30, "=B12*12", "تحويل"),
    ("صفقات مغلقة", "1–2", "—", "حسب التوافر"),
    ("نمو المتابعين", 1250, "=B14*12", "FB+IG+TikTok"),
]
for i, k in enumerate(kpis):
    r = 4 + i
    for c, v in enumerate(k, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = CENTER if c != 1 else RIGHT
        cell.font = LABEL_FONT if c == 1 else BODY_FONT
        if c == 1:
            cell.fill = ACCENT_FILL
        if c == 2 and isinstance(v, (int, float)):
            cell.font = INPUT_FONT
            cell.number_format = "#,##0"
        if c == 3 and str(v).startswith("="):
            cell.number_format = "#,##0"
set_widths(ws, [30, 18, 18, 24])

# ============================================================
# SHEET 8: MONTHLY TRACKING (fill-in)
# ============================================================
ws = wb.create_sheet("Monthly_Tracking")
ws.sheet_view.rightToLeft = True
title_row(ws, "متابعة شهرية للأداء (تُملأ شهرياً)", 9)
header_cells(ws, ["الشهر", "إنفاق (ج)", "Reach", "رسائل", "ليدات",
                  "ليدات مؤهلة", "تكلفة الليد (ج)", "صفقات", "ملاحظات"])
month_names = ["يونيو", "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر",
               "ديسمبر", "يناير", "فبراير", "مارس", "أبريل", "مايو"]
start = 4
for i, mn in enumerate(month_names):
    r = start + i
    ws.cell(row=r, column=1, value=mn).alignment = CENTER
    ws.cell(row=r, column=1).font = LABEL_FONT
    ws.cell(row=r, column=1).fill = GRAY_FILL
    for c in (2, 3, 4, 5, 6, 8):
        cell = ws.cell(row=r, column=c)
        cell.font = INPUT_FONT
        cell.alignment = CENTER
    # cost per lead formula
    cpl = ws.cell(row=r, column=7, value=f"=IFERROR(B{r}/E{r},0)")
    cpl.number_format = "0.0"
    cpl.alignment = CENTER
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).number_format = "#,##0" if c in (2, 3, 4, 5, 6, 8) else ws.cell(row=r, column=c).number_format

tot = start + len(month_names)
ws.cell(row=tot, column=1, value="الإجمالي").font = LABEL_FONT
for c in (2, 3, 4, 5, 6, 8):
    col = get_column_letter(c)
    ws.cell(row=tot, column=c, value=f"=SUM({col}{start}:{col}{tot-1})")
    ws.cell(row=tot, column=c).number_format = "#,##0"
ws.cell(row=tot, column=7, value=f"=IFERROR(B{tot}/E{tot},0)")
ws.cell(row=tot, column=7).number_format = "0.0"
for c in range(1, 10):
    cell = ws.cell(row=tot, column=c)
    cell.fill = ACCENT_FILL
    cell.font = LABEL_FONT
    cell.alignment = CENTER
    cell.border = BORDER
set_widths(ws, [14, 12, 12, 10, 10, 14, 16, 10, 24])

# ============================================================
# SHEET 9: LEAD FUNNEL
# ============================================================
ws = wb.create_sheet("Lead_Funnel")
ws.sheet_view.rightToLeft = True
title_row(ws, "قمع العميل وإدارة الليدات", 3)
header_cells(ws, ["المرحلة", "الوصف", "المسؤول"])
funnel = [
    ("1. الجذب", "إعلان / منشور يجذب الجمهور المستهدف", "Media Buyer + المحتوى"),
    ("2. التواصل", "العميل يبعت رسالة (ماسنجر/واتساب)", "الإعلان → البوت"),
    ("3. التأهيل", "البوت يسأل: الاهتمام + نوع الوحدة + الميزانية + وسيلة تواصل", "المساعد الذكي"),
    ("4. التحويل", "الليد المؤهل يتسجّل ويتحوّل للمبيعات", "Community Manager"),
    ("5. المتابعة", "اتصال → معاينة/Zoom → عرض", "متابعة المبيعات"),
    ("6. الإغلاق", "التعاقد + تحديث الحالة في الـ CRM", "متابعة المبيعات"),
    ("7. ما بعد البيع", "متابعة + إحالات + قناة إعادة البيع", "خدمة العملاء"),
]
for i, f in enumerate(funnel):
    r = 4 + i
    for c, v in enumerate(f, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = RIGHT if c == 2 else CENTER
        cell.font = LABEL_FONT if c == 1 else BODY_FONT
        if c == 1:
            cell.fill = ACCENT_FILL

# SLA note
sr = 4 + len(funnel) + 1
ws.cell(row=sr, column=1,
        value="⏱️ SLA مقترح: الليد المؤهل يتواصل معاه فريق المبيعات خلال ساعتين عمل كحد أقصى.")
ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=3)
ws.cell(row=sr, column=1).fill = GREEN_FILL
ws.cell(row=sr, column=1).font = LABEL_FONT
ws.cell(row=sr, column=1).alignment = RIGHT
ws.row_dimensions[sr].height = 28
set_widths(ws, [18, 56, 26])

# ============================================================
# SHEET 10: ROADMAP
# ============================================================
ws = wb.create_sheet("Roadmap")
ws.sheet_view.rightToLeft = True
title_row(ws, "خارطة الطريق السنوية (4 أرباع)", 3)
header_cells(ws, ["الربع", "التركيز", "أهم الأهداف"])
roadmap = [
    ("Q1 (شهور 1–3)", "التأسيس",
     "ضبط الهوية والقنوات · إطلاق حملات Meta · تركيب التتبّع والـ CRM · بدء تقدّم البناء · بيع 1–2 وحدة"),
    ("Q2 (شهور 4–6)", "التوسّع",
     "تحسين الإعلانات الفائزة · Retargeting + Lookalike · إطلاق TikTok · شهادات العملاء · موسم المغتربين"),
    ("Q3 (شهور 7–9)", "سلطة البراند",
     "Thought leadership · فيديوهات احترافية · Google/YouTube تجريبي · بدء تشويق Sky East"),
    ("Q4 (شهور 10–12)", "الحصاد والتوسعة",
     "إغلاق الوحدات المتبقية · إعلان المشروع القادم · بناء قائمة اهتمام · مراجعة سنوية وتخطيط السنة 2"),
]
for i, rm in enumerate(roadmap):
    r = 4 + i
    for c, v in enumerate(rm, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = RIGHT if c == 3 else CENTER
        cell.font = LABEL_FONT if c in (1, 2) else BODY_FONT
        if c == 1:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        if c == 2:
            cell.fill = ACCENT_FILL
    ws.row_dimensions[r].height = 46
set_widths(ws, [18, 18, 80])

# ---------- Save ----------
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Sky_Lines_Marketing_Plan_2026.xlsx")
wb.save(out)
print("Saved:", out)
print("Sheets:", wb.sheetnames)
