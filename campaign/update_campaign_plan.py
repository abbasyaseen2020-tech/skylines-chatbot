# -*- coding: utf-8 -*-
"""Update Campaign Plan: Faster Timeline + Engagement Boost"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reload existing workbook to update timeline + add engagement sheet
wb = load_workbook("D:/Apps/skylines-chatbot/campaign/SkyLines_Campaign_Plan_April2026.xlsx")

# Colors
HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
SUBHEADER_FILL = PatternFill("solid", start_color="6366F1")
SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HIGHLIGHT_FILL = PatternFill("solid", start_color="FEF3C7")
GREEN_FILL = PatternFill("solid", start_color="D1FAE5")
RED_FILL = PatternFill("solid", start_color="FEE2E2")
ORANGE_FILL = PatternFill("solid", start_color="FED7AA")
INPUT_FONT = Font(name="Arial", color="0000FF", size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="D1D5DB")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def style_subheader(cell):
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

# ============================================
# UPDATE TIMELINE SHEET (faster - drone is done!)
# ============================================
if "Timeline" in wb.sheetnames:
    del wb["Timeline"]

ws5 = wb.create_sheet("Timeline", 4)
ws5.sheet_view.rightToLeft = True

ws5["A1"] = "🚀 الجدول الزمني المُسرّع — الدرون اتعمل خلاص!"
ws5["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws5["A1"].fill = HEADER_FILL
ws5["A1"].alignment = CENTER
ws5.merge_cells("A1:D1")
ws5.row_dimensions[1].height = 30

headers = ["التاريخ", "اليوم", "المهمة", "المسؤول"]
for col_idx, h in enumerate(headers, start=1):
    cell = ws5.cell(row=3, column=col_idx, value=h)
    style_subheader(cell)

# UPDATED FAST TIMELINE - Drone is done
fast_timeline = [
    ("17 أبريل 2026", "الخميس (اليوم)", "✅ تشغيل Telegram Bot", "Claude — تم"),
    ("17 أبريل 2026", "الخميس", "✅ حذف بوستات التوظيف (HR + Architect)", "Claude — تم"),
    ("17 أبريل 2026", "الخميس", "✅ تجديد Railway سنة كاملة", "Abbas — تم"),
    ("17 أبريل 2026", "الخميس", "✅ تصوير الدرون", "Abbas — تم"),
    ("17 أبريل 2026", "الخميس مساءً", "🔄 رفع فيديو الدرون للمونتاج", "Abbas"),
    ("18 أبريل 2026", "الجمعة", "🎬 مونتاج فيديو الدرون (60 ثانية)", "فريق المونتاج"),
    ("18 أبريل 2026", "الجمعة", "🎨 تصميم 4 صور إعلانية (Carousel)", "Designer"),
    ("18 أبريل 2026", "الجمعة", "📝 مراجعة نصوص الإعلانات النهائية", "Abbas + Claude"),
    ("19 أبريل 2026", "السبت ⚡", "🚀 رفع الكامبين في Ads Manager (Draft)", "Claude"),
    ("19 أبريل 2026", "السبت", "✅ مراجعة Abbas + إجازة", "Abbas"),
    ("20 أبريل 2026", "الأحد 🔥", "▶️ بداية الحملة (Go Live)", "Facebook Ads"),
    ("20 أبريل 2026", "الأحد", "📊 مراقبة أول 6 ساعات", "Claude"),
    ("21 أبريل 2026", "الاثنين", "🔍 أول مراجعة (24 ساعة)", "Claude + Abbas"),
    ("23 أبريل 2026", "الأربعاء", "⚙️ مراجعة 3 أيام + تعديل Audiences", "Claude + Abbas"),
    ("27 أبريل 2026", "الأحد", "📈 تقرير الأسبوع الأول", "Claude"),
    ("4 مايو 2026", "الاثنين", "📈 تقرير الأسبوعين", "Claude"),
    ("20 مايو 2026", "الأربعاء", "📈 تقرير الشهر الأول", "Claude"),
    ("يومياً 9 م", "كل يوم", "📲 تقرير Telegram تلقائي", "البوت 🤖"),
]

for row_idx, row in enumerate(fast_timeline, start=4):
    for col_idx, value in enumerate(row, start=1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        cell.font = Font(name="Arial", size=11)
        if "🔥" in str(row[1]):
            cell.fill = GREEN_FILL
            cell.font = Font(name="Arial", bold=True, size=11)
        elif "⚡" in str(row[1]):
            cell.fill = ORANGE_FILL
        elif "✅" in str(row[2]):
            cell.fill = HIGHLIGHT_FILL

widths = [20, 20, 50, 25]
for i, w in enumerate(widths, start=1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# Add note about saved time
ws5.cell(row=24, column=1, value="🎉 وفّرنا 3 أيام كاملة — الإطلاق هيتم الأحد 20 أبريل بدلاً من الثلاثاء 22!")
ws5.cell(row=24, column=1).fill = GREEN_FILL
ws5.cell(row=24, column=1).font = Font(name="Arial", bold=True, size=12, color="065F46")
ws5.cell(row=24, column=1).alignment = CENTER
ws5.merge_cells("A24:D24")
ws5.row_dimensions[24].height = 30

# ============================================
# NEW SHEET: ENGAGEMENT BOOST STRATEGY
# ============================================
if "Engagement" in wb.sheetnames:
    del wb["Engagement"]

ws_eng = wb.create_sheet("Engagement")
ws_eng.sheet_view.rightToLeft = True

ws_eng["A1"] = "🔥 استراتيجية زيادة التفاعل (Engagement Boost)"
ws_eng["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws_eng["A1"].fill = HEADER_FILL
ws_eng["A1"].alignment = CENTER
ws_eng.merge_cells("A1:D1")
ws_eng.row_dimensions[1].height = 35

# Section 1: Current state
ws_eng["A3"] = "📊 الحالة الحالية للبيدج"
ws_eng["A3"].fill = SUBHEADER_FILL
ws_eng["A3"].font = SUBHEADER_FONT
ws_eng["A3"].alignment = CENTER
ws_eng.merge_cells("A3:D3")

current_state = [
    ["المؤشر", "القيمة", "المقارنة", "ملاحظة"],
    ["المتابعين", 17521, "كبير لبني سويف", "10x متوسط الشركات المنافسة"],
    ["Talking About", 207, "ضعيف جداً", "1.2% بس من المتابعين نشطين"],
    ["متوسط Engagement/بوست", 56, "ضعيف جداً", "العقارات 200-500"],
    ["Engagement Rate", "0.32%", "ضعيف", "السوق 2-5%"],
    ["إجمالي Likes (20 بوست)", 194, "ضعيف", "أقل من 10/بوست"],
    ["إجمالي Comments (20 بوست)", 201, "متوسط", "10/بوست"],
    ["إجمالي Shares (20 بوست)", 727, "ممتاز", "37/بوست — هنا قوتنا"],
]

for row_idx, row_data in enumerate(current_state, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_eng.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        if row_idx == 4:
            style_subheader(cell)
        else:
            cell.font = Font(name="Arial", size=11)
            if col_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=11)
                cell.fill = HIGHLIGHT_FILL

# Section 2: Engagement Tactics
section2_row = 13
ws_eng.cell(row=section2_row, column=1, value="🎯 10 تكتيكات لزيادة التفاعل").fill = SUBHEADER_FILL
ws_eng.cell(row=section2_row, column=1).font = SUBHEADER_FONT
ws_eng.cell(row=section2_row, column=1).alignment = CENTER
ws_eng.merge_cells(start_row=section2_row, start_column=1, end_row=section2_row, end_column=4)

tactics_headers = ["#", "التكتيك", "كيفية التنفيذ", "التأثير المتوقع"]
for col_idx, h in enumerate(tactics_headers, start=1):
    cell = ws_eng.cell(row=section2_row + 1, column=col_idx, value=h)
    style_subheader(cell)

tactics = [
    [1, "Comment Bait Posts",
     "اسأل سؤال في كل بوست: 'أنت تفضل ناصية ولا داخلي؟' / 'كم متر مناسب لعيلتك؟'",
     "+50-100% comments"],
    [2, "Pinned Post محدث",
     "ثبّت بوست العرض الذهبي (228 engagement) في أعلى البيدج كـ Featured Post",
     "+30% reach"],
    [3, "Reply لكل كومنت بسرعة",
     "البوت بيرد على كل كومنت في 30 ثانية تلقائياً (شغال الآن ✅)",
     "+200% comment threads"],
    [4, "Stories يومياً",
     "3-5 stories كل يوم: مشاهد من الموقع، before/after، فيديوهات الدرون",
     "+500% reach بدون فلوس"],
    [5, "Reels أسبوعياً",
     "5 Reels/أسبوع — قصير 15-30 ثانية — درون + transitions حديثة",
     "Reels reach 10x من البوستات"],
    [6, "Cross-Tag مع شركاء",
     "تاج العملاء/الموردين/المهندسين في كل بوست — يخلوا متابعينهم يشوفوا",
     "+20-40% organic reach"],
    [7, "Engagement Posts خفيفة",
     "بوست 'اختار التصميم اللي يعجبك' — 4 صور مع تصويت — مشاركة بسيطة",
     "+150% engagement"],
    [8, "Comment Pin",
     "ثبّت أحسن كومنت (سؤال أو شهادة) — يخلي الناس ترد عليه",
     "+80% comment depth"],
    [9, "Save-worthy Content",
     "بوست 'دليل شراء عقار في بني سويف' — checklist يحفظوه",
     "+300% saves (signal للـ algo)"],
    [10, "Carousel Posts",
     "بوست 4-10 صور (الوحدة، الموقع، التشطيب، السعر) — يطول وقت المشاهدة",
     "+150% time spent"],
]

for row_idx, row in enumerate(tactics, start=section2_row + 2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_eng.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.font = Font(name="Arial", size=10)
        if col_idx == 1:
            cell.font = Font(name="Arial", bold=True, size=11)
            cell.alignment = CENTER
            cell.fill = HIGHLIGHT_FILL
        if col_idx == 4:
            cell.fill = GREEN_FILL
            cell.font = Font(name="Arial", bold=True, size=10)

# Section 3: Posting Schedule
section3_row = section2_row + 14
ws_eng.cell(row=section3_row, column=1, value="📅 جدول النشر الأسبوعي المقترح").fill = SUBHEADER_FILL
ws_eng.cell(row=section3_row, column=1).font = SUBHEADER_FONT
ws_eng.cell(row=section3_row, column=1).alignment = CENTER
ws_eng.merge_cells(start_row=section3_row, start_column=1, end_row=section3_row, end_column=4)

schedule_headers = ["اليوم", "النوع", "الوقت", "المحتوى المقترح"]
for col_idx, h in enumerate(schedule_headers, start=1):
    cell = ws_eng.cell(row=section3_row + 1, column=col_idx, value=h)
    style_subheader(cell)

schedule = [
    ["السبت", "Reel + Story", "8 م + 12 م", "فيديو درون قصير + Story صور الموقع"],
    ["الأحد", "Carousel", "9 م", "4 صور — وحدات متاحة بأسعارها"],
    ["الاثنين", "Engagement Post", "8 م", "سؤال: 'حضرتك بتفضل سكني ولا تجاري؟'"],
    ["الثلاثاء", "Stories", "متعدد", "Behind the scenes — موقع البناء"],
    ["الأربعاء", "Reel", "9 م", "Reel transitions — قبل/بعد التشطيب"],
    ["الخميس", "Featured Post", "8 م", "العرض الحصري — الشقة عليك والخدمات علينا"],
    ["الجمعة", "Live أو Stories", "8 م", "بث مباشر — جولة في الموقع"],
]

for row_idx, row in enumerate(schedule, start=section3_row + 2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_eng.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        cell.font = Font(name="Arial", size=11)
        if col_idx == 1:
            cell.font = Font(name="Arial", bold=True, size=11)
            cell.fill = HIGHLIGHT_FILL

# Section 4: Quick Win Comments to Add
section4_row = section3_row + 11
ws_eng.cell(row=section4_row, column=1, value="💬 كومنتات Comment Bait جاهزة (انسخها لكل بوست)").fill = SUBHEADER_FILL
ws_eng.cell(row=section4_row, column=1).font = SUBHEADER_FONT
ws_eng.cell(row=section4_row, column=1).alignment = CENTER
ws_eng.merge_cells(start_row=section4_row, start_column=1, end_row=section4_row, end_column=4)

bait_comments = [
    ["#", "الكومنت", "نوع البوست المناسب", "متوقع رد"],
    [1, "إنت تفضل وحدة ناصية ولا داخلية؟ 🤔", "أي بوست وحدات", "20-50 رد"],
    [2, "كم متر مناسب لعيلة من 5 أفراد؟", "بوست المساحات", "30-80 رد"],
    [3, "أنت بتدور على سكن ولا استثمار؟", "بوست تعريفي", "40-100 رد"],
    [4, "إيه أكتر حاجة بتركز عليها لما تشتري شقة؟", "بوست تصميم", "30-70 رد"],
    [5, "تحب الموقع قرب الخدمات ولا في هدوء؟", "بوست الموقع", "20-60 رد"],
    [6, "أنت من بني سويف ولا مسافر؟", "Engagement post", "50-150 رد ⭐"],
    [7, "احكيلنا فين بتشتغل دلوقتي؟ (للمسافرين)", "بوست المغتربين", "100+ رد ⭐⭐"],
    [8, "كم مقدم مناسب لميزانيتك؟ 25%؟ 50%؟", "بوست السعر", "40-100 رد"],
    [9, "تفضل تشطيب جاهز ولا على العضم؟", "بوست التشطيب", "50-120 رد"],
    [10, "أيهم أهم: السعر، الموقع، التصميم، التقسيط؟", "أي بوست", "100+ رد ⭐⭐"],
]

for row_idx, row in enumerate(bait_comments, start=section4_row + 1):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_eng.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.font = Font(name="Arial", size=10)
        if row_idx == section4_row + 1:
            style_subheader(cell)
        elif col_idx == 1:
            cell.font = Font(name="Arial", bold=True, size=11)
            cell.fill = HIGHLIGHT_FILL
            cell.alignment = CENTER
        elif col_idx == 4 and "⭐" in str(value):
            cell.fill = GREEN_FILL
            cell.font = Font(name="Arial", bold=True, size=10)

# Section 5: Boost Posts (paid engagement)
section5_row = section4_row + 14
ws_eng.cell(row=section5_row, column=1, value="🚀 Boost Posts (تمويل التفاعل)").fill = SUBHEADER_FILL
ws_eng.cell(row=section5_row, column=1).font = SUBHEADER_FONT
ws_eng.cell(row=section5_row, column=1).alignment = CENTER
ws_eng.merge_cells(start_row=section5_row, start_column=1, end_row=section5_row, end_column=4)

boost_headers = ["البوست", "الميزانية", "المدة", "الهدف"]
for col_idx, h in enumerate(boost_headers, start=1):
    cell = ws_eng.cell(row=section5_row + 1, column=col_idx, value=h)
    style_subheader(cell)

boost_data = [
    ["بوست العرض الرئيسي (228 engagement حالياً)", "200 ج/يوم", "7 أيام", "Engagement → 1000+"],
    ["بوست الـ 161م² (147 engagement)", "100 ج/يوم", "7 أيام", "Engagement → 500+"],
    ["بوست 'وفر 200 ألف' (96 engagement)", "100 ج/يوم", "5 أيام", "Engagement → 400+"],
]

for row_idx, row in enumerate(boost_data, start=section5_row + 2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_eng.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        cell.font = Font(name="Arial", size=11)

# Total boost row
total_boost_row = section5_row + 2 + len(boost_data)
ws_eng.cell(row=total_boost_row, column=1, value="إجمالي الـ Boost (مرة واحدة)")
ws_eng.cell(row=total_boost_row, column=2, value="2,900 ج")
ws_eng.cell(row=total_boost_row, column=3, value="7 أيام")
ws_eng.cell(row=total_boost_row, column=4, value="منفصل عن ميزانية الـ 500/يوم")
for c in range(1, 5):
    ws_eng.cell(row=total_boost_row, column=c).fill = HIGHLIGHT_FILL
    ws_eng.cell(row=total_boost_row, column=c).font = Font(name="Arial", bold=True, size=11)
    ws_eng.cell(row=total_boost_row, column=c).border = BORDER
    ws_eng.cell(row=total_boost_row, column=c).alignment = CENTER

widths = [10, 35, 35, 25]
for i, w in enumerate(widths, start=1):
    ws_eng.column_dimensions[get_column_letter(i)].width = w

# ============================================
# UPDATE OVERVIEW with new launch date
# ============================================
ws1 = wb["Overview"]
# Find and update launch date
for row in ws1.iter_rows(min_row=4, max_row=20, min_col=1, max_col=2):
    if row[0].value == "تاريخ البدء":
        row[1].value = "الأحد 20 أبريل 2026 (تم تسريع 3 أيام)"
        row[1].fill = GREEN_FILL
        row[1].font = Font(name="Arial", bold=True, size=11)
    if row[0].value == "تاريخ المراجعة الأولى":
        row[1].value = "الاثنين 21 أبريل 2026"

wb.save("D:/Apps/skylines-chatbot/campaign/SkyLines_Campaign_Plan_April2026.xlsx")
print("✅ Plan updated with faster timeline + engagement strategies!")
