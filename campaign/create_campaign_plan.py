# -*- coding: utf-8 -*-
"""Generate Sky Lines Campaign Excel Plan"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
SUBHEADER_FILL = PatternFill("solid", start_color="6366F1")
SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HIGHLIGHT_FILL = PatternFill("solid", start_color="FEF3C7")
GREEN_FILL = PatternFill("solid", start_color="D1FAE5")
RED_FILL = PatternFill("solid", start_color="FEE2E2")
INPUT_FONT = Font(name="Arial", color="0000FF", size=11)
FORMULA_FONT = Font(name="Arial", color="000000", size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="D1D5DB")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def style_subheader(cell):
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

# ============================================
# SHEET 1: OVERVIEW
# ============================================
ws1 = wb.active
ws1.title = "Overview"
ws1.sheet_view.rightToLeft = True

ws1["A1"] = "خطة حملة Sky Lines الإعلانية — أبريل 2026"
ws1["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
ws1["A1"].fill = HEADER_FILL
ws1["A1"].alignment = CENTER
ws1.merge_cells("A1:F1")
ws1.row_dimensions[1].height = 40

overview_data = [
    ["العنصر", "القيمة"],
    ["اسم الحملة", "Sky Villas M7 — Expats Campaign"],
    ["المشروع", "Sky Villas M7 — شرق النيل، بني سويف"],
    ["العرض الحصري", "الشقة عليك والخدمات علينا (خصم لحد 200,000 ج)"],
    ["الميزانية اليومية", "500 جنيه"],
    ["الميزانية الشهرية", "15,000 جنيه"],
    ["تاريخ البدء", "الثلاثاء 22 أبريل 2026"],
    ["تاريخ المراجعة الأولى", "الجمعة 25 أبريل 2026"],
    ["هدف تكلفة الليد", "8-10 جنيه"],
    ["الليدات المتوقعة شهرياً", "1,500 - 1,800 ليد"],
    ["المنصات", "Facebook + Instagram + Messenger"],
    ["الأهداف الرئيسية", "Messages + Lead Forms"],
    ["عدد الوحدات المتاحة", "7 وحدات سكنية + 5 محلات تجارية"],
    ["AI Bot", "@SkyLinesLeadsBot + Messenger Bot"],
    ["تقارير التتبع", "Telegram يومياً 9 م"],
]

for row_idx, row_data in enumerate(overview_data, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        if row_idx == 3:
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
        else:
            if col_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=11)
                cell.fill = HIGHLIGHT_FILL
            else:
                cell.font = Font(name="Arial", size=11)

ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 50

# ============================================
# SHEET 2: BUDGET BREAKDOWN
# ============================================
ws2 = wb.create_sheet("Budget")
ws2.sheet_view.rightToLeft = True

ws2["A1"] = "توزيع الميزانية — 500 ج/يوم"
ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws2["A1"].fill = HEADER_FILL
ws2["A1"].alignment = CENTER
ws2.merge_cells("A1:G1")
ws2.row_dimensions[1].height = 30

headers = ["#", "الحملة", "الجمهور", "ميزانية يومية (ج)", "ميزانية شهرية (ج)", "نوع الحملة", "الهدف"]
for col_idx, h in enumerate(headers, start=1):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    style_subheader(cell)

campaigns = [
    [1, "Beni Suef Residents", "مقيمين بني سويف", 150, "Messages", "رسائل للبوت"],
    [2, "Expats — Education", "مسافرين درسوا في بني سويف", 125, "Messages", "رسائل للبوت"],
    [3, "Expats — Interests", "مسافرين مهتمين ببني سويف", 75, "Messages", "رسائل للبوت"],
    [4, "Lead Form Direct", "جميع المستهدفين — فورم مباشر", 100, "Lead Form", "فورم مباشر"],
    [5, "Retargeting Warm", "اللي تفاعل قبل كده", 50, "Messages", "تحويل"],
]

for row_idx, row in enumerate(campaigns, start=4):
    ws2.cell(row=row_idx, column=1, value=row[0]).border = BORDER
    ws2.cell(row=row_idx, column=2, value=row[1]).border = BORDER
    ws2.cell(row=row_idx, column=3, value=row[2]).border = BORDER
    ws2.cell(row=row_idx, column=4, value=row[3]).border = BORDER
    ws2.cell(row=row_idx, column=4).font = INPUT_FONT
    ws2.cell(row=row_idx, column=5, value=f"=D{row_idx}*30").border = BORDER
    ws2.cell(row=row_idx, column=5).font = FORMULA_FONT
    ws2.cell(row=row_idx, column=6, value=row[4]).border = BORDER
    ws2.cell(row=row_idx, column=7, value=row[5]).border = BORDER
    for c in range(1, 8):
        ws2.cell(row=row_idx, column=c).alignment = CENTER
        ws2.cell(row=row_idx, column=c).font = Font(name="Arial", size=11)
    ws2.cell(row=row_idx, column=4).font = INPUT_FONT
    ws2.cell(row=row_idx, column=5).font = FORMULA_FONT

# Total row
total_row = 4 + len(campaigns)
ws2.cell(row=total_row, column=1, value="").border = BORDER
ws2.cell(row=total_row, column=2, value="الإجمالي").font = Font(name="Arial", bold=True, size=12)
ws2.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})").font = Font(name="Arial", bold=True, color="000000")
ws2.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row-1})").font = Font(name="Arial", bold=True, color="000000")
for c in range(1, 8):
    ws2.cell(row=total_row, column=c).fill = HIGHLIGHT_FILL
    ws2.cell(row=total_row, column=c).border = BORDER
    ws2.cell(row=total_row, column=c).alignment = CENTER

# Number formats
for r in range(4, total_row + 1):
    ws2.cell(row=r, column=4).number_format = "#,##0"
    ws2.cell(row=r, column=5).number_format = "#,##0"

# Widths
widths = [5, 25, 30, 18, 18, 15, 25]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ROI section
ws2.cell(row=total_row + 3, column=1, value="العائد المتوقع").fill = HEADER_FILL
ws2.cell(row=total_row + 3, column=1).font = HEADER_FONT
ws2.merge_cells(start_row=total_row + 3, start_column=1, end_row=total_row + 3, end_column=7)
ws2.cell(row=total_row + 3, column=1).alignment = CENTER

roi_data = [
    ["البند", "القيمة (ج)", "ملاحظة"],
    ["استثمار الإعلانات الشهري", 15000, "Facebook Ads"],
    ["تكلفة تصوير الدرون", 3000, "مرة واحدة (مش شهري)"],
    ["إجمالي الاستثمار الشهر الأول", "=B" + str(total_row + 5) + "+B" + str(total_row + 6), "Ads + Drone"],
    ["سعر وحدة واحدة (متوسط)", 2200000, "Sky Villas M7"],
    ["لو بعنا وحدة واحدة فقط", "=B" + str(total_row + 8) + "-B" + str(total_row + 7), "صافي الربح"],
    ["ROI %", "=(B" + str(total_row + 9) + "/B" + str(total_row + 7) + ")", "عائد الاستثمار"],
]

for row_idx, row_data in enumerate(roi_data, start=total_row + 4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        if row_idx == total_row + 4:
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT

# Format ROI numbers
for r in range(total_row + 5, total_row + 10):
    ws2.cell(row=r, column=2).number_format = "#,##0"
ws2.cell(row=total_row + 10, column=2).number_format = "0.0%"

# ============================================
# SHEET 3: AUDIENCE TARGETING
# ============================================
ws3 = wb.create_sheet("Targeting")
ws3.sheet_view.rightToLeft = True

ws3["A1"] = "استهداف الجمهور — 3 Audiences"
ws3["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws3["A1"].fill = HEADER_FILL
ws3["A1"].alignment = CENTER
ws3.merge_cells("A1:C1")
ws3.row_dimensions[1].height = 30

audiences = [
    ("🔵 Audience 1: Beni Suef Residents (40%)", "200 ج/يوم"),
    ("Location", "بني سويف (40 km radius) + المقيمين فيها"),
    ("Age Range", "28-55"),
    ("Gender", "All"),
    ("Language", "Arabic"),
    ("Interests", "Real estate, Real estate investing, Property, Investment, Home improvement, Mortgage"),
    ("Behaviors", "Small business owners, Frequent travelers, Likely to move"),
    ("", ""),
    ("🟢 Audience 2: Expats by Education (25%)", "125 ج/يوم"),
    ("Location", "السعودية (الرياض، جدة، الدمام، الخبر) + الإمارات (دبي، أبوظبي) + الكويت + قطر + عُمان"),
    ("Age Range", "28-55"),
    ("Language", "Arabic (Egyptian)"),
    ("Schools", "جامعة بني سويف, Beni Suef University, كلية الهندسة جامعة بني سويف, كلية الطب جامعة بني سويف, Beni Suef Technical Institute, Beni Suef Azhar Institute, Ahmed Orabi Secondary Beni Suef, El Orman School Beni Suef, Beni Suef Experimental School"),
    ("Behaviors", "Expats — lived in Egypt, Expats — Egypt"),
    ("", ""),
    ("🟡 Audience 3: Expats by Interests (15%)", "75 ج/يوم"),
    ("Location", "نفس دول الخليج"),
    ("Interests", "Beni Suef, Beni Suef Governorate, محافظة بني سويف, شرق النيل بني سويف, El Wasta, إهناسيا, الفشن, نادي بني سويف الرياضي"),
    ("AND Also", "Real estate investing, Property investment, Egyptian real estate"),
    ("", ""),
    ("🔴 Audience 4: Lead Form (20%)", "100 ج/يوم"),
    ("Mix", "كل الـ Audiences السابقة بـ Lead Form Objective"),
    ("Form Fields", "الاسم, التليفون, المدينة الحالية, ميزانية الشراء"),
    ("", ""),
    ("🟣 Audience 5: Retargeting (10%)", "50 ج/يوم"),
    ("Custom Audience", "اللي تفاعل مع البيدج آخر 90 يوم + اللي بعتوا رسالة + Video Viewers 50%+"),
    ("Lookalike", "2% من قائمة العملاء الحاليين (لو متوفرة)"),
]

for row_idx, (label, value) in enumerate(audiences, start=3):
    ws3.cell(row=row_idx, column=1, value=label).border = BORDER
    ws3.cell(row=row_idx, column=2, value=value).border = BORDER
    if label.startswith(("🔵", "🟢", "🟡", "🔴", "🟣")):
        ws3.cell(row=row_idx, column=1).fill = SUBHEADER_FILL
        ws3.cell(row=row_idx, column=1).font = SUBHEADER_FONT
        ws3.cell(row=row_idx, column=2).fill = HIGHLIGHT_FILL
        ws3.cell(row=row_idx, column=2).font = Font(name="Arial", bold=True, size=11)
    else:
        ws3.cell(row=row_idx, column=1).font = Font(name="Arial", bold=True, size=11)
        ws3.cell(row=row_idx, column=2).font = Font(name="Arial", size=11)
    ws3.cell(row=row_idx, column=1).alignment = CENTER
    ws3.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 80

# ============================================
# SHEET 4: AD COPY
# ============================================
ws4 = wb.create_sheet("AdCopy")
ws4.sheet_view.rightToLeft = True

ws4["A1"] = "نصوص الإعلانات — 3 نسخ"
ws4["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws4["A1"].fill = HEADER_FILL
ws4["A1"].alignment = CENTER
ws4.merge_cells("A1:B1")
ws4.row_dimensions[1].height = 30

ads = [
    ("Ad 1: العرض الرئيسي (للمقيمين)", ""),
    ("Primary Text", """🎉 الشقة عليك والخدمات علينا!

عرض حصري من Sky Lines — وفّر لحد 200,000 جنيه على مشروع Sky Villas M7 في شرق النيل، بني سويف.

✅ مقدم 50% والباقي 15 شهر بدون فوايد
✅ موقع استراتيجي قرب طريق الأربع حارات
✅ تصميم كلاسيكي أوروبي فاخر
✅ 7 وحدات متاحة بس!

ابعتلنا رسالة دلوقتي ومساعدنا الذكي هيرد عليك فوراً 🤖"""),
    ("Headline", "وفّر 200,000 جنيه — Sky Villas M7"),
    ("Description", "عرض محدود — الشقة عليك والخدمات علينا"),
    ("CTA Button", "Send Message"),
    ("", ""),
    ("Ad 2: للمسافرين (Expats)", ""),
    ("Primary Text", """بعيد عن بني سويف بس قلبك لسه هناك؟ ❤️

لما ترجع تلاقي بيتك مستنيك في Sky Villas M7 — شرق النيل، قلب الحي الرابع.

🏡 شقق فاخرة بتصميم كلاسيكي أوروبي
💰 استثمار آمن بالجنيه المصري
📍 موقع مميز قرب كل الخدمات
💳 ادفع من الخليج بـ 50% مقدم والباقي قسط

🎁 عرض محدود: الخدمات علينا (توفير لحد 200 ألف ج)

ابعتلنا وهنرتب معاك كل حاجة عن بُعد 📲"""),
    ("Headline", "بيتك في بني سويف — جاهز استقبالك"),
    ("Description", "شقق Sky Villas M7 — استثمار بالجنيه"),
    ("CTA Button", "Send Message"),
    ("", ""),
    ("Ad 3: Urgency", ""),
    ("Primary Text", """⚠️ باقي 7 وحدات بس من Sky Villas M7!

آخر فرصة تحجز شقتك قبل ما يخلص العرض:
🔥 مقدم 50% فقط
🔥 قسط 15 شهر
🔥 خصم 200 ألف جنيه
🔥 استلام يوليو 2027

الأسعار في بني سويف بتطلع شهرياً — كل يوم تأخير بيزود السعر.

كلمنا دلوقتي قبل ما يخلص التوافر 📞"""),
    ("Headline", "آخر 7 وحدات — فرصة لن تتكرر"),
    ("CTA Button", "Send Message"),
]

for row_idx, (label, text) in enumerate(ads, start=3):
    ws4.cell(row=row_idx, column=1, value=label).border = BORDER
    ws4.cell(row=row_idx, column=2, value=text).border = BORDER
    if label.startswith("Ad"):
        ws4.cell(row=row_idx, column=1).fill = SUBHEADER_FILL
        ws4.cell(row=row_idx, column=1).font = SUBHEADER_FONT
        ws4.cell(row=row_idx, column=2).fill = HIGHLIGHT_FILL
    else:
        ws4.cell(row=row_idx, column=1).font = Font(name="Arial", bold=True, size=11)
        ws4.cell(row=row_idx, column=2).font = Font(name="Arial", size=11)
    ws4.cell(row=row_idx, column=1).alignment = CENTER
    ws4.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    # Adjust height for text-heavy rows
    if "Primary Text" in str(label):
        ws4.row_dimensions[row_idx].height = 200

ws4.column_dimensions["A"].width = 20
ws4.column_dimensions["B"].width = 80

# ============================================
# SHEET 5: TIMELINE
# ============================================
ws5 = wb.create_sheet("Timeline")
ws5.sheet_view.rightToLeft = True

ws5["A1"] = "الجدول الزمني — من اليوم للإطلاق"
ws5["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws5["A1"].fill = HEADER_FILL
ws5["A1"].alignment = CENTER
ws5.merge_cells("A1:D1")
ws5.row_dimensions[1].height = 30

headers = ["التاريخ", "اليوم", "المهمة", "المسؤول"]
for col_idx, h in enumerate(headers, start=1):
    cell = ws5.cell(row=3, column=col_idx, value=h)
    style_subheader(cell)

timeline = [
    ("17 أبريل 2026", "الخميس (اليوم)", "تجهيز النصوص والخطة", "Claude"),
    ("17 أبريل 2026", "الخميس", "تشغيل Telegram Bot للإشعارات", "Claude ✅"),
    ("17 أبريل 2026", "الخميس", "حذف بوستات التوظيف", "Abbas"),
    ("17 أبريل 2026", "الخميس", "تجديد Railway سنة كاملة", "Abbas ✅"),
    ("18 أبريل 2026", "الجمعة", "تجهيز الدرون والكاميرا", "فريق التصوير"),
    ("19 أبريل 2026", "السبت", "تصوير الدرون (4-6 ساعات)", "فريق التصوير"),
    ("20 أبريل 2026", "الأحد", "مونتاج فيديو الدرون (60 ثانية)", "فريق المونتاج"),
    ("20 أبريل 2026", "الأحد", "تصميم الصور الإعلانية (Carousel)", "Claude + Graphic Designer"),
    ("21 أبريل 2026", "الاثنين", "رفع الكامبين في Ads Manager", "Claude (Draft)"),
    ("21 أبريل 2026", "الاثنين", "مراجعة Abbas وإجازة", "Abbas"),
    ("22 أبريل 2026", "الثلاثاء 🚀", "بداية الحملة (Go Live)", "Facebook Ads"),
    ("23 أبريل 2026", "الأربعاء", "أول مراجعة (بعد 24 ساعة)", "Claude + Abbas"),
    ("25 أبريل 2026", "الجمعة", "مراجعة أول 3 أيام + تعديلات", "Claude + Abbas"),
    ("29 أبريل 2026", "الثلاثاء", "تقرير أسبوعي كامل", "Claude"),
    ("يومياً 9 م", "كل يوم", "تقرير Telegram تلقائي", "البوت 🤖"),
]

for row_idx, row in enumerate(timeline, start=4):
    for col_idx, value in enumerate(row, start=1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        cell.font = Font(name="Arial", size=11)
        if "🚀" in str(row[1]):
            cell.fill = GREEN_FILL
            cell.font = Font(name="Arial", bold=True, size=11)

widths = [20, 18, 50, 25]
for i, w in enumerate(widths, start=1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ============================================
# SHEET 6: KPIs / EXPECTED RESULTS
# ============================================
ws6 = wb.create_sheet("KPIs")
ws6.sheet_view.rightToLeft = True

ws6["A1"] = "المؤشرات والنتائج المتوقعة (KPIs)"
ws6["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws6["A1"].fill = HEADER_FILL
ws6["A1"].alignment = CENTER
ws6.merge_cells("A1:E1")
ws6.row_dimensions[1].height = 30

headers = ["المؤشر", "قبل الحملة", "الهدف الشهري", "الهدف (يومي)", "ملاحظات"]
for col_idx, h in enumerate(headers, start=1):
    cell = ws6.cell(row=3, column=col_idx, value=h)
    style_subheader(cell)

kpis = [
    ["Reach", "500", 100000, "=C4/30", "زيادة 200x"],
    ["Impressions", "2,000", 500000, "=C5/30", "زيادة 250x"],
    ["CTR % (Click-through)", "N/A", "2.5%", "", "معدل النقر"],
    ["Messages/Conversations", "5", 1800, "=C7/30", "هدف أساسي"],
    ["Lead Forms Submitted", "0", 600, "=C8/30", "هدف ثانوي"],
    ["Qualified Leads (من البوت)", "2", 450, "=C9/30", "25% qualification"],
    ["Cost per Lead (ج)", "—", 10, "", "HARD CAP — مش أكتر"],
    ["Cost per Qualified Lead (ج)", "—", 35, "", "هدف الـ CPQL"],
    ["Spend total (ج)", "0", 15000, "=C11/30", "Daily 500"],
    ["صفقات مغلقة (estimate)", "0-1", "5-15", "", "conversion 1-3%"],
]

for row_idx, row in enumerate(kpis, start=4):
    for col_idx, value in enumerate(row, start=1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        cell.font = Font(name="Arial", size=11)
        if col_idx == 1:
            cell.font = Font(name="Arial", bold=True, size=11)
            cell.fill = HIGHLIGHT_FILL
        # Input cells in blue
        if col_idx == 3 and isinstance(value, (int, float)):
            cell.font = Font(name="Arial", color="0000FF", size=11)
            cell.number_format = "#,##0"
        if col_idx == 4 and str(value).startswith("="):
            cell.number_format = "#,##0"

widths = [30, 15, 18, 15, 30]
for i, w in enumerate(widths, start=1):
    ws6.column_dimensions[get_column_letter(i)].width = w

# ============================================
# SHEET 7: CREATIVE SPECS
# ============================================
ws7 = wb.create_sheet("Creatives")
ws7.sheet_view.rightToLeft = True

ws7["A1"] = "مواصفات الإعلانات البصرية"
ws7["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws7["A1"].fill = HEADER_FILL
ws7["A1"].alignment = CENTER
ws7.merge_cells("A1:C1")
ws7.row_dimensions[1].height = 30

creative_data = [
    ("النوع", "المقاسات", "الملاحظات"),
    ("فيديو Feed (مربع)", "1080 × 1080 px — 15-60 ثانية", "للـ News Feed"),
    ("فيديو Stories/Reels", "1080 × 1920 px — 9:16 — حتى 90 ثانية", "للـ Stories والـ Reels"),
    ("صورة Feed", "1200 × 1200 px — PNG/JPG", "لـ Facebook Feed"),
    ("صورة Stories", "1080 × 1920 px — 9:16", "لـ Instagram Stories"),
    ("Carousel (4 صور)", "1080 × 1080 px لكل صورة", "لعرض 4 وحدات"),
    ("", "", ""),
    ("النص على الصورة", "لا يزيد عن 20%", "قاعدة فيسبوك"),
    ("ألوان Sky Lines", "بنفسجي فاخر (#5B21B6) + ذهبي (#F59E0B)", "الهوية البصرية"),
    ("Font عربي", "Cairo أو Tajawal أو IBM Plex Arabic", "خط حديث واضح"),
    ("Logo Placement", "أعلى يمين (10% من الصورة)", "دايماً يظهر"),
    ("CTA في التصميم", "زر واضح: 'احجز دلوقتي' أو 'كلمنا'", "لون مميز"),
    ("", "", ""),
    ("فيديو الدرون (60 ثانية)", "", ""),
    ("0-5 ثواني", "لقطة بعيدة للحي الرابع + النيل", "Establishing shot"),
    ("5-10 ثواني", "Zoom بطيء على المشروع", "Dramatic reveal"),
    ("10-20 ثواني", "دوران حول المشروع 360°", "إظهار الواجهة"),
    ("20-30 ثواني", "عرض الموقع + طريق الأربع حارات + المدرسة", "السياق"),
    ("30-40 ثواني", "دخول من فوق للداخل (لو ممكن)", "التفاصيل"),
    ("40-50 ثواني", "ارتفاع تدريجي — المدينة كلها", "Wide shot"),
    ("50-60 ثواني", "Logo + Call to Action", "الخاتمة"),
    ("", "", ""),
    ("النص الصوتي (Voice-over)", "", ""),
    ("اللغة", "عربي مصري احترافي", ""),
    ("النص", "بني سويف — المدينة اللي بتكبر كل يوم 🏙️ شرق النيل — قلب الحي الرابع. على بُعد دقائق من طريق الأربع حارات ومدرسة سان جورج. Sky Villas M7 — مشروع بتصميم كلاسيكي أوروبي. من سكاي لاينز للتطوير العقاري. 🎁 الشقة عليك والخدمات علينا — وفّر لحد 200 ألف جنيه. احجز دلوقتي — فرص محدودة", ""),
    ("الموسيقى", "cinematic/orchestral — حماسية", "تجنب المشهورة (copyright)"),
]

for row_idx, row in enumerate(creative_data, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = ws7.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.font = Font(name="Arial", size=11)
        if row_idx == 3:
            style_subheader(cell)
        elif col_idx == 1 and row[1] == "" and row[2] == "":
            cell.fill = HIGHLIGHT_FILL
            cell.font = Font(name="Arial", bold=True, size=12)
            cell.alignment = CENTER

widths = [30, 45, 35]
for i, w in enumerate(widths, start=1):
    ws7.column_dimensions[get_column_letter(i)].width = w

# ============================================
# SHEET 8: DAILY TRACKING
# ============================================
ws8 = wb.create_sheet("DailyTracking")
ws8.sheet_view.rightToLeft = True

ws8["A1"] = "تتبع يومي للأداء"
ws8["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws8["A1"].fill = HEADER_FILL
ws8["A1"].alignment = CENTER
ws8.merge_cells("A1:I1")
ws8.row_dimensions[1].height = 30

headers = ["التاريخ", "إنفاق (ج)", "Impressions", "Reach", "Clicks", "Messages", "Leads", "Cost/Lead (ج)", "ملاحظات"]
for col_idx, h in enumerate(headers, start=1):
    cell = ws8.cell(row=3, column=col_idx, value=h)
    style_subheader(cell)

# Pre-fill 30 days
from datetime import datetime, timedelta
start_date = datetime(2026, 4, 22)
for i in range(30):
    row_idx = 4 + i
    day = start_date + timedelta(days=i)
    ws8.cell(row=row_idx, column=1, value=day.strftime("%Y-%m-%d")).border = BORDER
    ws8.cell(row=row_idx, column=1).alignment = CENTER
    # Formula for cost per lead
    ws8.cell(row=row_idx, column=8, value=f"=IFERROR(B{row_idx}/G{row_idx},0)").border = BORDER
    ws8.cell(row=row_idx, column=8).number_format = "0.00"
    ws8.cell(row=row_idx, column=8).alignment = CENTER
    for c in [2, 3, 4, 5, 6, 7]:
        ws8.cell(row=row_idx, column=c).border = BORDER
        ws8.cell(row=row_idx, column=c).alignment = CENTER
        ws8.cell(row=row_idx, column=c).font = INPUT_FONT
    ws8.cell(row=row_idx, column=9).border = BORDER

# Totals row
total_row = 34
ws8.cell(row=total_row, column=1, value="الإجمالي").font = Font(name="Arial", bold=True, size=12)
ws8.cell(row=total_row, column=1).fill = HIGHLIGHT_FILL
for c in range(2, 8):
    col_letter = get_column_letter(c)
    ws8.cell(row=total_row, column=c, value=f"=SUM({col_letter}4:{col_letter}33)")
    ws8.cell(row=total_row, column=c).font = Font(name="Arial", bold=True, size=11)
    ws8.cell(row=total_row, column=c).fill = HIGHLIGHT_FILL
    ws8.cell(row=total_row, column=c).alignment = CENTER
    ws8.cell(row=total_row, column=c).border = BORDER
    ws8.cell(row=total_row, column=c).number_format = "#,##0"
ws8.cell(row=total_row, column=8, value=f"=IFERROR(B{total_row}/G{total_row},0)")
ws8.cell(row=total_row, column=8).font = Font(name="Arial", bold=True, size=11)
ws8.cell(row=total_row, column=8).fill = HIGHLIGHT_FILL
ws8.cell(row=total_row, column=8).alignment = CENTER
ws8.cell(row=total_row, column=8).border = BORDER
ws8.cell(row=total_row, column=8).number_format = "0.00"
ws8.cell(row=total_row, column=9).fill = HIGHLIGHT_FILL
ws8.cell(row=total_row, column=9).border = BORDER

widths = [15, 12, 15, 12, 10, 12, 10, 14, 25]
for i, w in enumerate(widths, start=1):
    ws8.column_dimensions[get_column_letter(i)].width = w

# Save
wb.save("D:/Apps/skylines-chatbot/campaign/SkyLines_Campaign_Plan_April2026.xlsx")
print("Excel saved successfully!")
